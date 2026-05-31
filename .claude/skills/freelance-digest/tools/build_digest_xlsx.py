"""Сборка XLSX фриланс-дайджеста из отфильтрованного JSON.

Детерминированный слой 3. Модель НЕ переписывает этот код — только готовит входной
JSON по схеме ниже и запускает скрипт.

Usage:
    python build_digest_xlsx.py <input.json> [output.xlsx]

Если output не задан — пишется рядом с input: <input-без-расширения>.xlsx

Схема входного JSON:
{
  "generated_at": "2026-04-21T19:00:00+03:00",   // ISO-время сборки
  "filters_applied": {
    "freshness_hours": 24,
    "min_budget_rub": 1000,
    "topics": ["AI", "лендинги", "парсинг", ...]
  },
  "sources_status": {                              // биржа -> статус строкой
    "fl.ru": "Работает: 9 заявок",
    "kwork.ru": "404 на все проекты"
  },
  "projects": [
    {
      "title": "AI-архитектор",
      "budget_rub": 50000,        // null если «по договорённости»
      "exchange": "fl.ru",
      "url": "https://www.fl.ru/projects/5500541/",
      "posted_at": "2026-04-16 13:07",
      "topic": "AI / Автоматизация"
    }
  ]
}

Лист 1 — заявки, сортировка по budget_rub desc (null уходит вниз).
Лист 2 — meta: фильтры + статусы бирж.
"""
import json
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

HEADERS = ["Название", "Бюджет (руб)", "Биржа", "Ссылка", "Время публикации", "Тема"]
WIDTHS = [60, 14, 14, 60, 24, 20]
HEADER_FILL = "2E86AB"
LINK_COLOR = "0563C1"


def build(src_path: Path, dst_path: Path) -> int:
    data = json.loads(src_path.read_text(encoding="utf-8"))
    projects = sorted(
        data.get("projects", []),
        key=lambda p: p.get("budget_rub") or 0,
        reverse=True,
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "digest"
    ws.append(HEADERS)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor=HEADER_FILL)
    for col_idx in range(1, len(HEADERS) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for p in projects:
        budget = p.get("budget_rub")
        ws.append([
            p.get("title", ""),
            budget if budget is not None else "по договорённости",
            p.get("exchange", ""),
            p.get("url", ""),
            p.get("posted_at", ""),
            p.get("topic", ""),
        ])
        url = p.get("url")
        if url:
            url_cell = ws.cell(row=ws.max_row, column=4)
            url_cell.hyperlink = url
            url_cell.font = Font(color=LINK_COLOR, underline="single")

    for i, w in enumerate(WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

    meta = wb.create_sheet("meta")
    meta.append(["Параметр", "Значение"])
    meta.append(["Дата сборки", data.get("generated_at", "")])
    f = data.get("filters_applied", {})
    meta.append(["Фильтр свежести (ч)", f.get("freshness_hours", "")])
    meta.append(["Мин. бюджет (руб)", f.get("min_budget_rub", "")])
    meta.append(["Темы", ", ".join(f.get("topics", []))])
    meta.append([])
    meta.append(["Биржа", "Статус"])
    for src, status in data.get("sources_status", {}).items():
        meta.append([src, status])
    meta.column_dimensions["A"].width = 30
    meta.column_dimensions["B"].width = 60

    dst_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(dst_path)
    return len(projects)


def main() -> None:
    if len(sys.argv) < 2:
        sys.exit("Usage: python build_digest_xlsx.py <input.json> [output.xlsx]")
    src = Path(sys.argv[1])
    if not src.exists():
        sys.exit(f"Нет входного файла: {src}")
    dst = Path(sys.argv[2]) if len(sys.argv) > 2 else src.with_suffix(".xlsx")
    rows = build(src, dst)
    print(f"OK: {dst}")
    print(f"Rows: {rows}")


if __name__ == "__main__":
    main()
