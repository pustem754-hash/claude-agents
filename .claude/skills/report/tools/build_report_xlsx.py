"""Generalized Excel (XLSX) report engine (data-driven).

Usage:
    python build_report_xlsx.py <spec.json> [out.xlsx]

Reads a JSON spec (see ../report_schema.md) and emits a styled .xlsx:
title row, source subtitle, header row, data rows, per-column number formats
and alignment, conditional formatting (sign: green >0 / red <0), color scales,
an optional Summary sheet, autofilter and freeze panes.

Data-driven by design: NEVER hardcode report content here. Everything comes
from the JSON spec. To change the look-and-feel, edit THIS engine once — do
not fork the script per report.
"""
from __future__ import annotations

import json
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# --- Palette ---
TITLE_FILL = "0B4D2F"      # deep green title band
HEADER_FILL = "1F4E79"     # blue header
GREEN_FILL = "C6EFCE"
GREEN_FONT = "006100"
RED_FILL = "FFC7CE"
RED_FONT = "9C0006"

_ALIGN_MAP = {
    "center": "center",
    "left": "left",
    "right": "right",
}


def _col_letter_for(spec_col, columns) -> str:
    """Resolve a spec `col` reference (letter like 'G' or 1-based index) to a letter."""
    if isinstance(spec_col, int):
        return get_column_letter(spec_col)
    s = str(spec_col).strip()
    if s.isdigit():
        return get_column_letter(int(s))
    return s.upper()


def build_xlsx(spec: dict, out_path: Path) -> Path:
    out_path.parent.mkdir(parents=True, exist_ok=True)

    title = spec.get("title", "Report")
    source = spec.get("source", "")
    sheet_name = spec.get("sheet_name", "Data")
    columns = spec.get("columns", [])
    rows = spec.get("rows", [])

    n_cols = max(1, len(columns))

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31] if sheet_name else "Data"

    title_font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor=HEADER_FILL)
    title_fill = PatternFill("solid", fgColor=TITLE_FILL)
    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")

    last_col_letter = get_column_letter(n_cols)

    # Row 1: title (merged across all columns)
    ws.merge_cells(f"A1:{last_col_letter}1")
    ws["A1"] = title
    ws["A1"].font = title_font
    ws["A1"].fill = title_fill
    ws["A1"].alignment = center
    ws.row_dimensions[1].height = 26

    # Row 2: source / subtitle
    has_source = bool(source)
    if has_source:
        ws.merge_cells(f"A2:{last_col_letter}2")
        ws["A2"] = source
        ws["A2"].font = Font(name="Calibri", size=9, italic=True, color="595959")
        ws["A2"].alignment = center
        ws.row_dimensions[2].height = 16
        header_row = 3
    else:
        header_row = 2

    # Header row
    for ci, col in enumerate(columns, start=1):
        name = col.get("name", "") if isinstance(col, dict) else str(col)
        c = ws.cell(row=header_row, column=ci, value=name)
        c.font = header_font
        c.fill = header_fill
        c.alignment = center
        c.border = border
    ws.row_dimensions[header_row].height = 22

    first_data_row = header_row + 1

    # Per-column alignment objects and number formats
    col_aligns = []
    col_formats = []
    for col in columns:
        if isinstance(col, dict):
            align = _ALIGN_MAP.get(str(col.get("align", "left")).lower(), "left")
            col_formats.append(col.get("number_format"))
        else:
            align = "left"
            col_formats.append(None)
        col_aligns.append(Alignment(horizontal=align, vertical="center"))

    # Data rows
    for ri, row in enumerate(rows):
        excel_row = first_data_row + ri
        for ci in range(1, n_cols + 1):
            value = row[ci - 1] if ci - 1 < len(row) else None
            c = ws.cell(row=excel_row, column=ci, value=value)
            c.border = border
            c.font = Font(name="Calibri", size=10)
            c.alignment = col_aligns[ci - 1]
            fmt = col_formats[ci - 1]
            if fmt:
                c.number_format = fmt

    last_row = first_data_row + len(rows) - 1 if rows else header_row

    # Conditional formatting (sign): green >0 / red <0
    for rule in spec.get("conditional", []):
        if not isinstance(rule, dict):
            continue
        if rule.get("type") == "sign" and rows:
            letter = _col_letter_for(rule.get("col"), columns)
            rng = f"{letter}{first_data_row}:{letter}{last_row}"
            ws.conditional_formatting.add(rng, CellIsRule(
                operator="greaterThan", formula=["0"], stopIfTrue=False,
                fill=PatternFill("solid", fgColor=GREEN_FILL), font=Font(color=GREEN_FONT)))
            ws.conditional_formatting.add(rng, CellIsRule(
                operator="lessThan", formula=["0"], stopIfTrue=False,
                fill=PatternFill("solid", fgColor=RED_FILL), font=Font(color=RED_FONT)))

    # Color scales (gradient by magnitude)
    for spec_col in spec.get("colorscale", []):
        if not rows:
            break
        letter = _col_letter_for(spec_col, columns)
        rng = f"{letter}{first_data_row}:{letter}{last_row}"
        ws.conditional_formatting.add(rng, ColorScaleRule(
            start_type="min", start_color="FFFFFF",
            mid_type="percentile", mid_value=50, mid_color="BDD7EE",
            end_type="max", end_color=HEADER_FILL))

    # Column widths
    for ci, col in enumerate(columns, start=1):
        width = col.get("width") if isinstance(col, dict) else None
        if width:
            ws.column_dimensions[get_column_letter(ci)].width = width

    # Freeze panes
    freeze = spec.get("freeze")
    if freeze:
        ws.freeze_panes = freeze
    else:
        ws.freeze_panes = f"A{first_data_row}"

    # Autofilter (over header + data range)
    if spec.get("autofilter", False):
        ws.auto_filter.ref = f"A{header_row}:{last_col_letter}{max(last_row, header_row)}"

    # Optional Summary sheet
    summary = spec.get("summary", [])
    if summary:
        ws2 = wb.create_sheet("Summary")
        ws2["A1"] = "Summary"
        ws2["A1"].font = Font(bold=True, size=12)
        for i, pair in enumerate(summary, start=3):
            if not pair:
                continue
            key = pair[0] if len(pair) > 0 else ""
            val = pair[1] if len(pair) > 1 else ""
            ws2.cell(row=i, column=1, value=key).font = Font(bold=True)
            ws2.cell(row=i, column=2, value=val)
        ws2.column_dimensions["A"].width = 32
        ws2.column_dimensions["B"].width = 30

    wb.save(out_path)
    return out_path


def main(argv) -> int:
    if len(argv) < 2:
        print("Usage: python build_report_xlsx.py <spec.json> [out.xlsx]")
        return 2

    spec_path = Path(argv[1])
    if not spec_path.exists():
        print(f"ERROR: spec not found: {spec_path}")
        return 1

    spec = json.loads(spec_path.read_text(encoding="utf-8"))
    if "columns" not in spec:
        print("ERROR: spec must contain a 'columns' array")
        return 1

    if len(argv) >= 3:
        out_path = Path(argv[2])
    else:
        out_path = spec_path.with_suffix(".xlsx")

    built = build_xlsx(spec, out_path)
    n_rows = len(spec.get("rows", []))
    n_cols = len(spec.get("columns", []))
    print(f"OK: {built}")
    print(f"columns: {n_cols}  rows: {n_rows}  size: {built.stat().st_size} bytes")
    return 0


if __name__ == "__main__":
    raise SystemExit(main(sys.argv))
