import json
import statistics
import io
from datetime import date

# ── Data loading ──────────────────────────────────────────────────────────────
with open(r"C:\Users\Пользователь\OneDrive\Документы\claude-agents\agent-runtime\shared\hh-vacancies.json",
          encoding="utf-8") as f:
    data = json.load(f)

vacancies = data["vacancies"]
PARSED_AT  = data["parsed_at"]
AREA       = data["search_area"]
TOTAL      = data["total_vacancies"]

# ── Classification helpers ────────────────────────────────────────────────────
BACKEND_KW  = ["backend", "бэкенд", "python", "java", "php", "django",
                "back-разработчик", "ml-инженер", "it-разработчик"]
FRONTEND_KW = ["frontend", "фронтенд", "фронт", "react", "js/react", "front-end"]
FULL_KW     = ["fullstack", "фулстек", "универсальный it", "full stack"]

def classify(name: str) -> str:
    n = name.lower()
    if any(k in n for k in FULL_KW):
        return "Fullstack"
    if any(k in n for k in FRONTEND_KW):
        return "Frontend"
    if any(k in n for k in BACKEND_KW):
        return "Backend"
    return "Другое"

for v in vacancies:
    v["category"] = classify(v["name"])

# ── Salary analytics (RUR only) ───────────────────────────────────────────────
rur_vacs = [v for v in vacancies if v["salary_currency"] == "RUR"]

def midpoint(v):
    """Return midpoint salary; use whichever bound is available."""
    lo, hi = v["salary_from"], v["salary_to"]
    if lo and hi:
        return (lo + hi) / 2
    return lo or hi

rur_mids = [midpoint(v) for v in rur_vacs if midpoint(v)]

with_salary    = [v for v in vacancies if v["salary"] is not None]
without_salary = [v for v in vacancies if v["salary"] is None]
eur_vacs       = [v for v in vacancies if v["salary_currency"] == "EUR"]

# Category salary stats (RUR midpoints)
def cat_stats(cat):
    mids = [midpoint(v) for v in rur_vacs if v["category"] == cat and midpoint(v)]
    if not mids:
        return None
    return {
        "count": len(mids),
        "min":   int(min(mids)),
        "max":   int(max(mids)),
        "mean":  int(statistics.mean(mids)),
        "median": int(statistics.median(mids)),
    }

backend_stats  = cat_stats("Backend")
frontend_stats = cat_stats("Frontend")

# Employer counts
employer_counts = {}
for v in vacancies:
    employer_counts[v["employer"]] = employer_counts.get(v["employer"], 0) + 1
top_employers = sorted(employer_counts.items(), key=lambda x: -x[1])[:8]

# Category distribution
from collections import Counter
cat_dist = Counter(v["category"] for v in vacancies)

# ── Overall RUR stats ─────────────────────────────────────────────────────────
overall_stats = None
if rur_mids:
    overall_stats = {
        "count":  len(rur_mids),
        "min":    int(min(rur_mids)),
        "max":    int(max(rur_mids)),
        "mean":   int(statistics.mean(rur_mids)),
        "median": int(statistics.median(rur_mids)),
    }

# ═════════════════════════════════════════════════════════════════════════════
# XLSX REPORT
# ═════════════════════════════════════════════════════════════════════════════
import openpyxl
from openpyxl.styles import (PatternFill, Font, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference, PieChart

wb = openpyxl.Workbook()

# ── Helpers ────────────────────────────────────────────────────────────────
HDR_FILL   = PatternFill("solid", fgColor="1F4E79")
HDR_FONT   = Font(bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(bold=True, size=14, color="1F4E79")
SUB_FONT   = Font(bold=True, size=11, color="2E75B6")
THIN       = Side(style="thin", color="BFBFBF")
BORDER     = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CAT_COLORS = {"Backend": "D6E4F0", "Frontend": "D5F5E3",
              "Fullstack": "FDEBD0", "Другое": "F5EEF8"}

def set_hdr(ws, row, col, text):
    c = ws.cell(row=row, column=col, value=text)
    c.fill   = HDR_FILL
    c.font   = HDR_FONT
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = BORDER
    return c

def data_cell(ws, row, col, value, fmt=None, fill_hex=None, bold=False):
    c = ws.cell(row=row, column=col, value=value)
    c.alignment = Alignment(vertical="center", wrap_text=True)
    c.border    = BORDER
    if fmt:
        c.number_format = fmt
    if fill_hex:
        c.fill = PatternFill("solid", fgColor=fill_hex)
    if bold:
        c.font = Font(bold=True)
    return c

def fmt_rub(n):
    if n is None:
        return "—"
    return f"{int(n):,}".replace(",", " ") + " ₽"

# ── Sheet 1: All vacancies ─────────────────────────────────────────────────
ws1 = wb.active
ws1.title = "Все вакансии"

ws1.row_dimensions[1].height = 20
ws1.row_dimensions[2].height = 16
ws1.row_dimensions[3].height = 30

ws1.merge_cells("A1:H1")
t = ws1.cell(1, 1, "Вакансии разработчиков — Москва, HH.ru")
t.font      = Font(bold=True, size=16, color="1F4E79")
t.alignment = Alignment(horizontal="center", vertical="center")

ws1.merge_cells("A2:H2")
s = ws1.cell(2, 1, f"Дата парсинга: {PARSED_AT}  |  Всего вакансий: {TOTAL}")
s.font      = Font(size=10, italic=True, color="595959")
s.alignment = Alignment(horizontal="center", vertical="center")

headers = ["№", "Название вакансии", "Работодатель",
           "Категория", "Зарплата (указана)", "От (руб.)", "До (руб.)", "Валюта"]
col_widths = [5, 42, 30, 12, 32, 14, 14, 10]

for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
    set_hdr(ws1, 3, ci, h)
    ws1.column_dimensions[get_column_letter(ci)].width = w

for ri, v in enumerate(vacancies, 1):
    row = ri + 3
    ws1.row_dimensions[row].height = 20
    cat   = v["category"]
    fill  = CAT_COLORS.get(cat, "FFFFFF")
    sal_d = v["salary"] if v["salary"] else "Не указана"

    data_cell(ws1, row, 1, ri,             fill_hex=fill)
    data_cell(ws1, row, 2, v["name"],      fill_hex=fill)
    data_cell(ws1, row, 3, v["employer"],  fill_hex=fill)
    data_cell(ws1, row, 4, cat,            fill_hex=fill)
    data_cell(ws1, row, 5, sal_d,          fill_hex=fill)
    data_cell(ws1, row, 6, v["salary_from"], fmt="#,##0", fill_hex=fill)
    data_cell(ws1, row, 7, v["salary_to"],   fmt="#,##0", fill_hex=fill)
    data_cell(ws1, row, 8, v["salary_currency"] or "—", fill_hex=fill)

ws1.freeze_panes = "A4"

# ── Sheet 2: Analytics ────────────────────────────────────────────────────
ws2 = wb.create_sheet("Аналитика")

def section_title(ws, row, col, text, span=4):
    ws.merge_cells(start_row=row, start_column=col,
                   end_row=row,   end_column=col + span - 1)
    c = ws.cell(row=row, column=col, value=text)
    c.font      = Font(bold=True, size=13, color="FFFFFF")
    c.fill      = PatternFill("solid", fgColor="2E75B6")
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row].height = 22
    return c

def kv(ws, row, key, val, key_fill="E8F0FE", val_fill="FFFFFF"):
    c1 = ws.cell(row=row, column=1, value=key)
    c1.font      = Font(bold=True, size=10)
    c1.fill      = PatternFill("solid", fgColor=key_fill)
    c1.alignment = Alignment(vertical="center", indent=1)
    c1.border    = BORDER
    ws.row_dimensions[row].height = 18

    c2 = ws.cell(row=row, column=2, value=val)
    c2.fill      = PatternFill("solid", fgColor=val_fill)
    c2.alignment = Alignment(vertical="center", indent=1)
    c2.border    = BORDER
    return c1, c2

ws2.column_dimensions["A"].width = 32
ws2.column_dimensions["B"].width = 22
ws2.column_dimensions["C"].width = 5
ws2.column_dimensions["D"].width = 24
ws2.column_dimensions["E"].width = 18
ws2.column_dimensions["F"].width = 18
ws2.column_dimensions["G"].width = 18

r = 1
# ── Block 1: Overview ──────────────────────────────────────────────────────
section_title(ws2, r, 1, "Общая статистика")
r += 1
kv(ws2, r, "Всего вакансий",              TOTAL)
r += 1
kv(ws2, r, "С указанной зарплатой",       len(with_salary))
r += 1
kv(ws2, r, "Без указанной зарплаты",      len(without_salary))
r += 1
kv(ws2, r, "Доля с указанной зарплатой",  f"{len(with_salary)/TOTAL*100:.1f}%")
r += 1
kv(ws2, r, "Вакансии в EUR",              len(eur_vacs))
r += 1
kv(ws2, r, "Вакансии в RUR",              len(rur_vacs))
r += 2

# ── Block 2: Salary stats (RUR) ───────────────────────────────────────────
section_title(ws2, r, 1, "Статистика зарплат в рублях (RUR)")
r += 1
if overall_stats:
    kv(ws2, r, "Количество вакансий с RUR",   overall_stats["count"])
    r += 1
    kv(ws2, r, "Минимальная зарплата",         fmt_rub(overall_stats["min"]))
    r += 1
    kv(ws2, r, "Максимальная зарплата",        fmt_rub(overall_stats["max"]))
    r += 1
    kv(ws2, r, "Средняя зарплата",             fmt_rub(overall_stats["mean"]))
    r += 1
    kv(ws2, r, "Медианная зарплата",           fmt_rub(overall_stats["median"]))
    r += 1
else:
    r += 1
r += 1

# ── Block 3: Category distribution ───────────────────────────────────────
section_title(ws2, r, 1, "Распределение по категориям")
r += 1
hdr_row = r
for ci, h in enumerate(["Категория", "Вакансий", "Доля"], 1):
    set_hdr(ws2, r, ci, h)
r += 1
for cat, cnt in sorted(cat_dist.items(), key=lambda x: -x[1]):
    fill = CAT_COLORS.get(cat, "FFFFFF")
    data_cell(ws2, r, 1, cat, fill_hex=fill)
    data_cell(ws2, r, 2, cnt, fill_hex=fill)
    data_cell(ws2, r, 3, f"{cnt/TOTAL*100:.1f}%", fill_hex=fill)
    r += 1
r += 1

# ── Block 4: Category salary comparison ──────────────────────────────────
section_title(ws2, r, 1, "Зарплаты по категориям (RUR)", span=5)
r += 1
for ci, h in enumerate(["Категория", "Вак. с зарплатой",
                         "Мин.", "Макс.", "Среднее", "Медиана"], 1):
    set_hdr(ws2, r, ci, h)
r += 1
for cat in ["Backend", "Frontend", "Fullstack", "Другое"]:
    st = cat_stats(cat)
    fill = CAT_COLORS.get(cat, "FFFFFF")
    if st:
        data_cell(ws2, r, 1, cat,                fill_hex=fill)
        data_cell(ws2, r, 2, st["count"],         fill_hex=fill)
        data_cell(ws2, r, 3, fmt_rub(st["min"]),  fill_hex=fill)
        data_cell(ws2, r, 4, fmt_rub(st["max"]),  fill_hex=fill)
        data_cell(ws2, r, 5, fmt_rub(st["mean"]), fill_hex=fill)
        data_cell(ws2, r, 6, fmt_rub(st["median"]),fill_hex=fill)
    else:
        data_cell(ws2, r, 1, cat, fill_hex=fill)
        for ci in range(2, 7):
            data_cell(ws2, r, ci, "—", fill_hex=fill)
    r += 1
r += 1

# ── Block 5: Top employers ────────────────────────────────────────────────
section_title(ws2, r, 1, "Топ работодателей")
r += 1
for ci, h in enumerate(["Работодатель", "Вакансий"], 1):
    set_hdr(ws2, r, ci, h)
r += 1
for emp, cnt in top_employers:
    data_cell(ws2, r, 1, emp)
    data_cell(ws2, r, 2, cnt)
    r += 1

ws2.freeze_panes = "A2"

# ── Sheet 3: Charts data ──────────────────────────────────────────────────
ws3 = wb.create_sheet("Графики")
ws3.column_dimensions["A"].width = 20
ws3.column_dimensions["B"].width = 15

# Category distribution data for chart
ws3["A1"] = "Категория"
ws3["B1"] = "Количество"
ws3["A1"].font = Font(bold=True)
ws3["B1"].font = Font(bold=True)
for i, (cat, cnt) in enumerate(sorted(cat_dist.items(), key=lambda x: -x[1]), 2):
    ws3[f"A{i}"] = cat
    ws3[f"B{i}"] = cnt

# Pie chart: category distribution
pie = PieChart()
pie.title    = "Распределение вакансий по категориям"
pie.style    = 10
pie.height   = 12
pie.width    = 18

labels = Reference(ws3, min_col=1, min_row=2, max_row=1+len(cat_dist))
data   = Reference(ws3, min_col=2, min_row=1, max_row=1+len(cat_dist))
pie.add_data(data, titles_from_data=True)
pie.set_categories(labels)
ws3.add_chart(pie, "D1")

# Bar chart: salary by category
ws3["A8"]  = "Категория"
ws3["B8"]  = "Среднее (RUR)"
ws3["C8"]  = "Медиана (RUR)"
ws3["A8"].font = Font(bold=True)
ws3["B8"].font = Font(bold=True)
ws3["C8"].font = Font(bold=True)
bar_cats = ["Backend", "Frontend", "Fullstack", "Другое"]
for i, cat in enumerate(bar_cats, 9):
    st = cat_stats(cat)
    ws3.cell(i, 1, cat)
    ws3.cell(i, 2, st["mean"]   if st else 0)
    ws3.cell(i, 3, st["median"] if st else 0)

bar = BarChart()
bar.type     = "col"
bar.style    = 10
bar.title    = "Средняя и медианная зарплата по категориям (RUR)"
bar.y_axis.title = "Зарплата, руб."
bar.x_axis.title = "Категория"
bar.height   = 12
bar.width    = 18

data_b = Reference(ws3, min_col=2, max_col=3, min_row=8, max_row=12)
cats_b = Reference(ws3, min_col=1, min_row=9, max_row=12)
bar.add_data(data_b, titles_from_data=True)
bar.set_categories(cats_b)
ws3.add_chart(bar, "D17")

XLSX_PATH = (r"C:\Users\Пользователь\OneDrive\Документы\claude-agents"
             r"\agent-runtime\outputs\hh_vacancies_report.xlsx")
wb.save(XLSX_PATH)
print(f"XLSX saved: {XLSX_PATH}")

# ═════════════════════════════════════════════════════════════════════════════
# PDF REPORT
# ═════════════════════════════════════════════════════════════════════════════
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import cm
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer, Table,
                                TableStyle, HRFlowable, PageBreak)
from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_RIGHT
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
import os, sys

# ── Font registration ─────────────────────────────────────────────────────
# Try to register a Cyrillic-capable font; fall back to built-in Helvetica.
FONT_DIRS = [
    r"C:\Windows\Fonts",
    "/usr/share/fonts/truetype/dejavu",
    "/usr/share/fonts",
]
FONT_PAIRS = [
    ("DejaVuSans",       "DejaVuSans.ttf",       None),
    ("DejaVuSans",       "DejaVuSans.ttf",       None),
    ("Arial",            "arial.ttf",             "arialbd.ttf"),
    ("Calibri",          "calibri.ttf",           "calibrib.ttf"),
    ("TimesNewRoman",    "times.ttf",             "timesbd.ttf"),
]

BASE_FONT  = "Helvetica"
BOLD_FONT  = "Helvetica-Bold"
REGISTERED = False

for font_dir in FONT_DIRS:
    if REGISTERED:
        break
    for fname, regular, bold_file in FONT_PAIRS:
        reg_path  = os.path.join(font_dir, regular)
        if not os.path.exists(reg_path):
            continue
        try:
            pdfmetrics.registerFont(TTFont(fname, reg_path))
            BASE_FONT = fname
            if bold_file:
                bold_path = os.path.join(font_dir, bold_file)
                if os.path.exists(bold_path):
                    pdfmetrics.registerFont(TTFont(fname + "-Bold", bold_path))
                    BOLD_FONT = fname + "-Bold"
                else:
                    BOLD_FONT = fname
            else:
                BOLD_FONT = fname
            REGISTERED = True
            print(f"Font registered: {fname} from {font_dir}")
            break
        except Exception as e:
            print(f"Font {fname} failed: {e}", file=sys.stderr)

if not REGISTERED:
    print("WARNING: No Cyrillic font found; Cyrillic text may not render.",
          file=sys.stderr)

# ── Colour palette ────────────────────────────────────────────────────────
C_DARK  = colors.HexColor("#1F4E79")   # dark blue
C_MID   = colors.HexColor("#2E75B6")   # mid blue
C_LIGHT = colors.HexColor("#D6E4F0")   # pale blue
C_ACCENT= colors.HexColor("#E8F5E9")   # pale green
C_GRAY  = colors.HexColor("#F5F5F5")   # light gray
C_RED   = colors.HexColor("#C0392B")
C_TEXT  = colors.HexColor("#212121")
C_SUB   = colors.HexColor("#555555")

W, H = A4
M_L, M_R, M_T, M_B = 2*cm, 2*cm, 2.5*cm, 2*cm
USABLE = W - M_L - M_R

# ── Styles ────────────────────────────────────────────────────────────────
styles = getSampleStyleSheet()

def ps(name, parent="Normal", **kw):
    kw.setdefault("fontName", BASE_FONT)
    return ParagraphStyle(name, parent=styles[parent], **kw)

S_TITLE  = ps("title",  fontSize=22, leading=28, textColor=C_DARK,
               fontName=BOLD_FONT, alignment=TA_CENTER, spaceAfter=6)
S_SUB    = ps("sub",    fontSize=11, leading=14, textColor=C_SUB,
               alignment=TA_CENTER, spaceAfter=4)
S_H1     = ps("h1",     fontSize=14, leading=18, textColor=colors.white,
               fontName=BOLD_FONT, spaceAfter=4, spaceBefore=12)
S_H2     = ps("h2",     fontSize=11, leading=14, textColor=C_DARK,
               fontName=BOLD_FONT, spaceAfter=4, spaceBefore=8)
S_BODY   = ps("body",   fontSize=9,  leading=13, textColor=C_TEXT, spaceAfter=2)
S_SMALL  = ps("small",  fontSize=8,  leading=11, textColor=C_SUB)
S_METRIC = ps("metric", fontSize=18, leading=22, textColor=C_DARK,
               fontName=BOLD_FONT, alignment=TA_CENTER)
S_MLBL   = ps("mlbl",   fontSize=8,  leading=10, textColor=C_SUB,
               alignment=TA_CENTER)

# ── Helper: section header ────────────────────────────────────────────────
def section_header(text):
    tbl = Table([[Paragraph(text, S_H1)]], colWidths=[USABLE])
    tbl.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,-1), C_DARK),
        ("TOPPADDING",    (0,0), (-1,-1), 7),
        ("BOTTOMPADDING", (0,0), (-1,-1), 7),
        ("LEFTPADDING",   (0,0), (-1,-1), 10),
        ("RIGHTPADDING",  (0,0), (-1,-1), 10),
        ("ROWBACKGROUNDS",(0,0), (-1,-1), [C_DARK]),
    ]))
    return tbl

# ── Helper: metric card row ───────────────────────────────────────────────
def metric_cards(items):
    """items: list of (label, value) tuples, max 4."""
    n   = len(items)
    cw  = USABLE / n
    hdr = [[Paragraph(v, S_METRIC) for _, v in items]]
    lbl = [[Paragraph(l, S_MLBL)   for l, _ in items]]
    tbl = Table(hdr + lbl, colWidths=[cw]*n)
    tbl.setStyle(TableStyle([
        ("BACKGROUND",    (0,0), (-1,-1), C_LIGHT),
        ("ROWBACKGROUNDS",(0,1), (-1,-1), [C_ACCENT]),
        ("TOPPADDING",    (0,0), (-1,-1), 8),
        ("BOTTOMPADDING", (0,0), (-1,-1), 8),
        ("LEFTPADDING",   (0,0), (-1,-1), 4),
        ("RIGHTPADDING",  (0,0), (-1,-1), 4),
        ("GRID",          (0,0), (-1,-1), 0.5, colors.white),
        ("ALIGN",         (0,0), (-1,-1), "CENTER"),
        ("VALIGN",        (0,0), (-1,-1), "MIDDLE"),
    ]))
    return tbl

# ── Helper: styled table ──────────────────────────────────────────────────
def styled_table(rows, col_widths, hdr_bg=C_DARK):
    tbl = Table(rows, colWidths=col_widths, repeatRows=1)
    style = [
        ("BACKGROUND",    (0,0), (-1, 0), hdr_bg),
        ("TEXTCOLOR",     (0,0), (-1, 0), colors.white),
        ("FONTNAME",      (0,0), (-1, 0), BOLD_FONT),
        ("FONTSIZE",      (0,0), (-1, 0), 9),
        ("ALIGN",         (0,0), (-1, 0), "CENTER"),
        ("VALIGN",        (0,0), (-1,-1), "MIDDLE"),
        ("FONTNAME",      (0,1), (-1,-1), BASE_FONT),
        ("FONTSIZE",      (0,1), (-1,-1), 8),
        ("ROWBACKGROUNDS",(0,1), (-1,-1), [colors.white, C_GRAY]),
        ("GRID",          (0,0), (-1,-1), 0.4, colors.HexColor("#BDBDBD")),
        ("TOPPADDING",    (0,0), (-1,-1), 5),
        ("BOTTOMPADDING", (0,0), (-1,-1), 5),
        ("LEFTPADDING",   (0,0), (-1,-1), 6),
        ("RIGHTPADDING",  (0,0), (-1,-1), 6),
    ]
    tbl.setStyle(TableStyle(style))
    return tbl

# ── Build story ───────────────────────────────────────────────────────────
PDF_PATH = (r"C:\Users\Пользователь\OneDrive\Документы\claude-agents"
            r"\agent-runtime\outputs\hh_vacancies_report.pdf")

doc = SimpleDocTemplate(
    PDF_PATH,
    pagesize=A4,
    leftMargin=M_L, rightMargin=M_R,
    topMargin=M_T, bottomMargin=M_B,
    title="Отчёт по вакансиям разработчиков — HH.ru",
    author="report-generator agent",
)

story = []

# ── Cover / header ────────────────────────────────────────────────────────
story.append(Spacer(1, 1.2*cm))
story.append(Paragraph("Рынок труда разработчиков", S_TITLE))
story.append(Paragraph("Москва · HH.ru · 10 апреля 2026", S_SUB))
story.append(HRFlowable(width=USABLE, thickness=2, color=C_DARK, spaceAfter=10))
story.append(Spacer(1, 0.3*cm))

# ── Key metrics ───────────────────────────────────────────────────────────
story.append(metric_cards([
    ("Всего вакансий",           str(TOTAL)),
    ("С указанной зарплатой",    str(len(with_salary))),
    ("Без зарплаты",             str(len(without_salary))),
    ("Вакансии в EUR",           str(len(eur_vacs))),
]))
story.append(Spacer(1, 0.5*cm))

if overall_stats:
    story.append(metric_cards([
        ("Мин. зарплата (RUR)",   fmt_rub(overall_stats["min"])),
        ("Макс. зарплата (RUR)",  fmt_rub(overall_stats["max"])),
        ("Средняя (RUR)",         fmt_rub(overall_stats["mean"])),
        ("Медиана (RUR)",         fmt_rub(overall_stats["median"])),
    ]))
    story.append(Spacer(1, 0.5*cm))

# ── Key findings ──────────────────────────────────────────────────────────
story.append(section_header("Ключевые выводы"))
story.append(Spacer(1, 0.2*cm))

pct_no_sal = len(without_salary) / TOTAL * 100
backend_pct  = cat_dist.get("Backend", 0) / TOTAL * 100
frontend_pct = cat_dist.get("Frontend", 0) / TOTAL * 100

findings = [
    f"<b>{pct_no_sal:.0f}% вакансий ({len(without_salary)} из {TOTAL})</b> не содержат указания зарплаты — "
    f"работодатели предпочитают обсуждать условия на собеседовании.",

    f"<b>Backend</b> доминирует: {cat_dist.get('Backend',0)} вакансий "
    f"({backend_pct:.0f}%), <b>Frontend</b> — {cat_dist.get('Frontend',0)} "
    f"({frontend_pct:.0f}%). Направление Fullstack представлено слабо.",

    f"Зарплатный диапазон в рублях охватывает значительный разброс: "
    f"от <b>{fmt_rub(overall_stats['min'])}</b> (Junior-уровень) до "
    f"<b>{fmt_rub(overall_stats['max'])}</b> (Senior / Lead). "
    f"Медианное значение — <b>{fmt_rub(overall_stats['median'])}</b>." if overall_stats else "",

    f"Работодатель <b>AbeloHost B.V.</b> предлагает зарплаты в EUR "
    f"(1 800–4 000 €/мес.), что существенно выше рублёвого рынка.",

    f"Компания <b>Айтисокет</b> размещает наибольшее число вакансий — 2 позиции "
    f"(Backend и Frontend), аналогично <b>AbeloHost B.V.</b> и "
    f"<b>Институту экономики и развития транспорта</b>.",
]
for f in findings:
    if f:
        story.append(Paragraph(f"• {f}", S_BODY))
        story.append(Spacer(1, 0.15*cm))

story.append(Spacer(1, 0.4*cm))

# ── Distribution by category ──────────────────────────────────────────────
story.append(section_header("Распределение вакансий по категориям"))
story.append(Spacer(1, 0.2*cm))

cat_rows = [["Категория", "Вакансий", "Доля, %"]]
for cat, cnt in sorted(cat_dist.items(), key=lambda x: -x[1]):
    cat_rows.append([cat, str(cnt), f"{cnt/TOTAL*100:.1f}%"])

story.append(styled_table(cat_rows, [USABLE*0.5, USABLE*0.25, USABLE*0.25]))
story.append(Spacer(1, 0.5*cm))

# ── Salary statistics (RUR) ───────────────────────────────────────────────
story.append(section_header("Статистика зарплат по категориям (RUR)"))
story.append(Spacer(1, 0.2*cm))

sal_rows = [["Категория", "Вак. с зарп.", "Мин.", "Макс.", "Среднее", "Медиана"]]
for cat in ["Backend", "Frontend", "Fullstack", "Другое"]:
    st = cat_stats(cat)
    if st:
        sal_rows.append([
            cat,
            str(st["count"]),
            fmt_rub(st["min"]),
            fmt_rub(st["max"]),
            fmt_rub(st["mean"]),
            fmt_rub(st["median"]),
        ])
    else:
        sal_rows.append([cat, "—", "—", "—", "—", "—"])

cws = [USABLE*0.15, USABLE*0.12, USABLE*0.18, USABLE*0.18, USABLE*0.19, USABLE*0.18]
story.append(styled_table(sal_rows, cws))
story.append(Spacer(1, 0.3*cm))
story.append(Paragraph(
    "* Зарплатные вилки рассчитаны как среднее от (salary_from + salary_to) / 2. "
    "Для вакансий только с нижней или верхней границей используется соответствующее значение.",
    S_SMALL))
story.append(Spacer(1, 0.5*cm))

# ── Salary with/without ───────────────────────────────────────────────────
story.append(section_header("Наличие зарплаты в объявлениях"))
story.append(Spacer(1, 0.2*cm))

sw_rows = [["Статус", "Количество", "Доля"]]
sw_rows.append(["Зарплата указана",    str(len(with_salary)),
                f"{len(with_salary)/TOTAL*100:.1f}%"])
sw_rows.append(["Зарплата не указана", str(len(without_salary)),
                f"{len(without_salary)/TOTAL*100:.1f}%"])
story.append(styled_table(sw_rows, [USABLE*0.5, USABLE*0.25, USABLE*0.25]))
story.append(Spacer(1, 0.5*cm))

# ── Top employers ─────────────────────────────────────────────────────────
story.append(section_header("Топ работодателей"))
story.append(Spacer(1, 0.2*cm))

emp_rows = [["#", "Работодатель", "Вакансий"]]
for i, (emp, cnt) in enumerate(top_employers, 1):
    emp_rows.append([str(i), emp, str(cnt)])
story.append(styled_table(emp_rows, [USABLE*0.08, USABLE*0.72, USABLE*0.20]))
story.append(Spacer(1, 0.5*cm))

# ── Full vacancy table ────────────────────────────────────────────────────
story.append(PageBreak())
story.append(section_header("Полный список вакансий"))
story.append(Spacer(1, 0.2*cm))

vac_rows = [["№", "Название", "Работодатель", "Кат.", "Зарплата"]]
for i, v in enumerate(vacancies, 1):
    sal_str = v["salary"] if v["salary"] else "Не указана"
    vac_rows.append([
        str(i),
        v["name"],
        v["employer"],
        v["category"],
        sal_str,
    ])

cws2 = [USABLE*0.05, USABLE*0.33, USABLE*0.25, USABLE*0.09, USABLE*0.28]
tbl2 = Table(vac_rows, colWidths=cws2, repeatRows=1)

cat_color_map = {
    "Backend":   colors.HexColor("#D6E4F0"),
    "Frontend":  colors.HexColor("#D5F5E3"),
    "Fullstack": colors.HexColor("#FDEBD0"),
    "Другое":    colors.HexColor("#F5EEF8"),
}

tbl2_style = [
    ("BACKGROUND",    (0,0), (-1, 0), C_DARK),
    ("TEXTCOLOR",     (0,0), (-1, 0), colors.white),
    ("FONTNAME",      (0,0), (-1, 0), BOLD_FONT),
    ("FONTSIZE",      (0,0), (-1, 0), 8),
    ("ALIGN",         (0,0), (-1, 0), "CENTER"),
    ("VALIGN",        (0,0), (-1,-1), "MIDDLE"),
    ("FONTNAME",      (0,1), (-1,-1), BASE_FONT),
    ("FONTSIZE",      (0,1), (-1,-1), 7.5),
    ("GRID",          (0,0), (-1,-1), 0.4, colors.HexColor("#BDBDBD")),
    ("TOPPADDING",    (0,0), (-1,-1), 4),
    ("BOTTOMPADDING", (0,0), (-1,-1), 4),
    ("LEFTPADDING",   (0,0), (-1,-1), 5),
    ("RIGHTPADDING",  (0,0), (-1,-1), 5),
]

# Colour data rows by category
for i, v in enumerate(vacancies, 1):
    bg = cat_color_map.get(v["category"], colors.white)
    tbl2_style.append(("BACKGROUND", (0, i), (-1, i), bg))
    # Dim "not specified" salary text
    if not v["salary"]:
        tbl2_style.append(("TEXTCOLOR", (4, i), (4, i), colors.HexColor("#AAAAAA")))

tbl2.setStyle(TableStyle(tbl2_style))
story.append(tbl2)

story.append(Spacer(1, 0.5*cm))
story.append(HRFlowable(width=USABLE, thickness=1, color=C_MID))
story.append(Spacer(1, 0.15*cm))
story.append(Paragraph(
    f"Отчёт сгенерирован автоматически агентом report-generator · "
    f"Источник: api.hh.ru · Дата: {PARSED_AT}",
    S_SMALL))

# ── Build PDF ─────────────────────────────────────────────────────────────
doc.build(story)
print(f"PDF saved: {PDF_PATH}")
